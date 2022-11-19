from __future__ import annotations

from contextlib import contextmanager
from pathlib import Path
from typing import NamedTuple

from Bio import SeqIO
from Bio.SeqIO.FastaIO import SimpleFastaParser
from openpyxl import load_workbook

from .types import Container, Type


class Sequence(NamedTuple):
    id: str
    seq: str
    extras: dict[str, str] = dict()

    _tr_normalize = str.maketrans('?', 'N', '-')

    def normalize(self):
        return Sequence(self.id, self.seq.translate(self._tr_normalize).upper(), self.extras)


class Sequences(Container[Sequence]):
    @classmethod
    def fromFile(cls, file: SequenceFile, *args, **kwargs) -> Sequences:
        return cls(file.read, *args, **kwargs)

    def normalize(self) -> Sequences:
        return Sequences(lambda: (seq.normalize() for seq in self))


class SequenceFile(Type):
    """Handlers for sequence files"""

    def __init__(self, path: Path):
        self.path = path

    def read(self, *args, **kwargs) -> iter[Sequence]:
        raise NotImplementedError()

    def write(self, distances: iter[Sequence], *args, **kwargs) -> None:
        raise NotImplementedError()


class SequenceFileHandler:
    def __init__(self, file: SequenceFile):
        self.coroutine = file.iter_write()
        next(self.coroutine)

    def write(self, sequence: Sequence) -> None:
        self.coroutine.send(sequence)

    def close(self) -> None:
        self.coroutine.close()


class Fasta(SequenceFile):
    def read(self) -> iter[Sequence]:
        with open(self.path, 'r') as handle:
            for data in SimpleFastaParser(handle):
                yield Sequence(*data)


class Genbank(SequenceFile):
    def read(self) -> iter[Sequence]:
        # Bio.GenBank.Scanner
        for data in SeqIO.parse(self.path, 'genbank'):
            yield Sequence(data.id, data.seq)


class Tabular(SequenceFile):
    def iter_rows(self):
        raise NotImplementedError()

    def iter_write(self):
        raise NotImplementedError()

    @contextmanager
    def open(self, mode='r'):
        if mode == 'r':
            yield self.iter_rows()
        elif mode == 'w':
            handler = SequenceFileHandler(self)
            yield handler
            handler.close()
        else:
            raise Exception('Unknown mode')

    def read(
        self,
        idHeader: str = None,
        seqHeader: str = None,
        hasHeader: bool = False,
        idColumn: int = 0,
        seqColumn: int = 1,
    ) -> iter[Sequence]:

        with self.open() as rows:

            # Checking headers
            if idHeader and seqHeader:
                hasHeader = True
            if hasHeader:
                headers = next(rows)
                idColumn, seqColumn = headers.index(idHeader), headers.index(seqHeader)

            # Getting id and seq
            for row in rows:
                if len(row) <= 1:
                    continue
                id = row[idColumn]
                seq = row[seqColumn]

                extras = {}
                if hasHeader:
                    extras = {
                        k: v for c, (k, v) in enumerate(zip(headers, row))
                        if c not in (idColumn, seqColumn)
                    }

                yield Sequence(id, seq, extras)

    def getHeader(self):
        with open(self.path) as file:
            print(file.readline())


class Tabfile(Tabular, SequenceFile):
    def iter_rows(self) -> iter[tuple[str, ...]]:
        with open(self.path) as file:
            for line in file:
                yield line.strip().split('\t')

    def getHeader(self):
        with open(self.path) as file:
            return file.readline().strip().split('\t')

    def iter_write(self, idHeader='seqid', seqHeader='sequence'):
        with open(self.path, 'w') as file:
            try:
                sequence = yield
                extraHeaders = sequence.extras.keys()
                extraHeadersString = '\t'.join(extraHeaders)
                file.write(f'{idHeader}\t{extraHeadersString}\t{seqHeader}\n')
                while True:
                    extras = sequence.extras.values()
                    extrasString = '\t'.join(extras)
                    file.write(f'{sequence.id}\t{extrasString}\t{sequence.seq}\n')
                    sequence = yield  # Yield expression
            except GeneratorExit:
                return


class Excel(Tabular, SequenceFile):
    def iter_rows(self) -> iter[tuple[str, ...]]:
        wb = load_workbook(filename=self.path, read_only=True)
        ws = wb.worksheets[0]
        for row in ws.iter_rows(values_only=True):
            row = list(row)
            while row and row[-1] is None:
                del row[-1]
            yield [x if x else '' for x in row]
        wb.close
