from __future__ import annotations

from contextlib import contextmanager
from pathlib import Path
from typing import NamedTuple, Generator, Generic, TypeVar
from itertools import chain
from abc import ABC, abstractmethod
from openpyxl import load_workbook

from .types import Type

Item = TypeVar('Item')
Row = tuple[str, ...]


class _FileHandlerMeta(type(ABC), type(Type)):
    pass


class FileHandler(ABC, Type, Generic[Item], metaclass=_FileHandlerMeta):
    """
    Abstract interface for reading/writing items to a file. Mimics io.IOBase.
    Subclasses must implement _iter_read() and _iter_write(), as well as
    _open_readable() and _open_writable() if necessary.

    Usage examples:

        with FileHandler('my_file.txt') as file:
            for item in file:
                print(item)

        with FileHandler('my_file.txt', 'w') as file:
            for item in my_items:
                file.write(item)

    """

    def __init__(self, *args, **kwargs):
        self._open(*args, **kwargs)
        next(self.it)

    def __enter__(self):
        return self

    def __exit__(self, type, value, tb):
        self.close()

    def __iter__(self):
        assert self.readable()
        return self

    def __next__(self):
        assert self.readable()
        return next(self.it)

    def _open(self, path: Path, mode: 'r' | 'w' = 'r', *args, **kwargs):
        self.path = path
        self.mode = mode
        if mode == 'r':
            self._open_readable(*args, **kwargs)
        elif mode == 'w':
            self._open_writable(*args, **kwargs)
        else:
            raise ValueError('Mode must be "r" or "w"')
        self.closed = False

    def _open_readable(self, *args, **kwargs):
        self.it = self._iter_read()

    def _open_writable(self, *args, **kwargs):
        self.it = self._iter_write()

    @abstractmethod
    def _iter_read(self) -> iter[Item]:
        yield  # ready
        while False:
            yield Item()

    @abstractmethod
    def _iter_write(self) -> Generator[None, Item, None]:
        try:
            while True:
                _ = yield
        except GeneratorExit:
            return

    def close(self):
        self.it.close()
        self.closed = True

    def read(self) -> Item | None:
        try:
            return next(self.it)
        except StopIteration:
            return None

    def write(self, item: Item) -> None:
        self.it.send(item)

    def readable(self) -> bool:
        return self.mode == 'r'

    def writable(self) -> bool:
        return self.mode == 'w'


class Tabular(FileHandler):
    def _open_readable(
        self,
        columns: iter[int | str] = None,
        has_headers: bool = False,
        get_all_columns: bool = False,
    ):
        if columns is not None:
            columns = tuple(columns)
            if not len(columns):
                raise ValueError('Columns argument must contain at least one item')
            if isinstance(columns[0], str):
                has_headers = True
        self.columns = columns
        self.has_headers = has_headers
        self.get_all_columns = get_all_columns
        self._header_row = None
        self._column_order = None
        super()._open_readable()

    def _iter_read(self) -> iter[Row]:
        rows = self._iter_read_rows()
        if self.has_headers:
            try:
                self._header_row = next(rows)
            except StopIteration:
                self._header_row = None
                yield  # ready
                return
        if self.columns is None:
            yield  # ready
            yield from rows
        else:
            yield from self._iter_columns(rows)

    def _iter_columns(self, rows: iter[Row]) -> iter[Row]:
        columns = self.columns
        if isinstance(columns[0], str):
            try:
                columns = tuple(self._header_row.index(x) for x in columns)
            except Exception as e:
                missing = set(columns) - set(self._header_row)
                raise ValueError(f'Column header(s) not found in file: {missing}') from e
        if self.get_all_columns:
            if self.has_headers:
                first_row = self._header_row
            else:
                first_row = next(rows)
                rows = chain([first_row], rows)
            extra_columns = set(range(len(first_row))) - set(columns)
            columns = columns + tuple(extra_columns)
        self._column_order = columns

        yield  # ready
        for row in rows:
            yield tuple(row[x] for x in columns)

    def _open_writable(
        self,
        columns: iter[str] = None,
    ):
        if columns is not None:
            columns = tuple(columns)
            if not len(columns):
                raise ValueError('Columns argument must contain at least one item')
        self.columns = columns
        super()._open_writable()

    def _iter_write(self) -> Generator[None, Row, None]:
        rows = self._iter_write_rows()
        next(rows)
        if self.columns is not None:
            rows.send(self.columns)
        try:
            while True:
                row = yield
                rows.send(row)
        except GeneratorExit:
            return

    @property
    def headers(self) -> Row | None:
        assert self.readable()
        if not self.has_headers:
            return None
        if self._column_order:
            return tuple(self._header_row[x] for x in self._column_order)
        return self._header_row

    @classmethod
    def get_headers(cls, path: Path) -> Row:
        with cls(path) as handler:
            return handler.read()

    @abstractmethod
    def _iter_read_rows(self) -> iter[Row]:
        while False:
            yield Row()

    @abstractmethod
    def _iter_write_rows(self) -> Generator[None, Row, None]:
        try:
            while True:
                _ = yield
        except GeneratorExit:
            return


class Tabfile(Tabular, FileHandler):
    def _iter_read_rows(self) -> iter[Row]:
        with open(self.path, 'r') as file:
            for line in file:
                line = line[:-1]
                if not line:
                    break
                yield tuple(line.split('\t'))

    def _iter_write_rows(self) -> Generator[None, Row, None]:
        with open(self.path, 'w') as file:
            try:
                while True:
                    row = yield
                    text = '\t'.join(row)
                    file.write(text + '\n')
            except GeneratorExit:
                return


class Excel(Tabular, FileHandler):
    def _iter_read_rows(self) -> iter[Row]:
        wb = load_workbook(filename=self.path, read_only=True)
        try:
            ws = wb.worksheets[0]
            for row in ws.iter_rows(values_only=True):
                row = list(row)
                while row and row[-1] is None:
                    del row[-1]
                if not row:
                    break
                yield tuple(x if x else '' for x in row)
        finally:
            wb.close()

    def _iter_write_rows(self) -> Generator[None, Row, None]:
        raise NotImplementedError()
