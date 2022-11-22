from __future__ import annotations

from contextlib import contextmanager
from pathlib import Path
from typing import NamedTuple, Generator
from itertools import chain

from openpyxl import load_workbook

from .types import Type

Row = tuple[str, ...]


class ReadHandler:
    def __init__(
        self,
        path: Path,
        columns: iter[int | str] = None,
        has_headers: bool = False,
        get_all_columns: bool = False,
    ):
        if columns is not None:
            columns = tuple(columns)
            if not len(columns):
                raise ValueError('Columns argument must contain at least one item')
            if isinstance(columns[0], str):
                has_headers = True
        self.path = path
        self.columns = columns
        self.has_headers = has_headers
        self.get_all_columns = get_all_columns
        self._header_row = None
        self._column_order = None
        self.it = self.iter_read()
        next(self.it)  # get ready

    def __enter__(self):
        return self

    def __exit__(self, type, value, tb):
        self.close()

    def __iter__(self):
        return self

    def __next__(self):
        return next(self.it)

    def close(self) -> None:
        self.it.close()

    def _iter_read_rows(self) -> iter[Row]:
        raise NotImplementedError()

    def iter_read(self) -> iter[Row]:
        rows = self._iter_read_rows()
        if self.has_headers:
            self._header_row = next(rows)
        if self.columns is None:
            yield  # ready
            yield from rows
        else:
            yield from self._iter_columns(rows)

    def _iter_columns(self, rows: iter[Row]) -> iter[Row]:
        columns = self.columns
        if isinstance(columns[0], str):
            try:
                columns = tuple(self._header_row.index(x) for x in columns)
            except Exception as e:
                missing = set(columns) - set(self._header_row)
                raise ValueError(f'Column header(s) not found in file: {missing}') from e
        if self.get_all_columns:
            if self.has_headers:
                first_row = self._header_row
            else:
                first_row = next(rows)
                rows = chain([first_row], rows)
            extra_columns = set(range(len(first_row))) - set(columns)
            columns = columns + tuple(extra_columns)
        self._column_order = columns

        yield  # ready
        for row in rows:
            yield tuple(row[x] for x in columns)

    def read(self) -> Row | None:
        try:
            return next(self.it)
        except StopIteration:
            return None

    @property
    def headers(self) -> Row | None:
        if not self.has_headers:
            return None
        if self._column_order:
            return tuple(self._header_row[x] for x in self._column_order)
        return self._header_row


class WriteHandler:
    def __init__(
        self,
        path: Path,
        columns: iter[str] = None,
    ):
        if columns is not None:
            columns = tuple(columns)
            if not len(columns):
                raise ValueError('Columns argument must contain at least one item')
        self.path = path
        self.columns = columns
        self.gen = self.gen_write()
        next(self.gen)

    def __enter__(self):
        return self

    def __exit__(self, type, value, tb):
        self.close()

    def close(self) -> None:
        self.gen.close()

    def _gen_write_rows(self) -> Generator[None, Row, None]:
        raise NotImplementedError()

    def gen_write(self) -> Generator[None, Row, None]:
        rows = self._gen_write_rows()
        next(rows)
        if self.columns is not None:
            rows.send(self.columns)
        try:
            while True:
                row = yield
                rows.send(row)
        except GeneratorExit:
            return

    def write(self, row: Row) -> None:
        self.gen.send(row)


class Tabular(Type):
    read_handler = ReadHandler
    write_handler = WriteHandler

    @classmethod
    def open(cls, path: Path, mode: 'r' | 'w' = 'r', *args, **kwargs) -> ReadHandler | WriteHandler:
        if mode == 'r':
            return cls.read_handler(path, *args, **kwargs)
        elif mode == 'w':
            return cls.write_handler(path, *args, **kwargs)
        raise ValueError('Mode must be "r" or "w"')

    @classmethod
    def headers(cls, path: Path) -> Row:
        with cls.read_handler(path) as handler:
            return handler.read()


class TabfileReadHandler(ReadHandler):
    def _iter_read_rows(self) -> iter[Row]:
        with open(self.path, 'r') as file:
            for line in file:
                line = line[:-1]
                if not line:
                    break
                yield tuple(line.split('\t'))


class TabfileWriteHandler(WriteHandler):
    def _gen_write_rows(self) -> Generator[None, Row, None]:
        with open(self.path, 'w') as file:
            try:
                while True:
                    row = yield
                    text = '\t'.join(row)
                    file.write(text + '\n')
            except GeneratorExit:
                return


class Tabfile(Tabular):
    read_handler = TabfileReadHandler
    write_handler = TabfileWriteHandler


class ExcelReadHandler(ReadHandler):
    def _iter_read_rows(self) -> iter[Row]:
        wb = load_workbook(filename=self.path, read_only=True)
        try:
            ws = wb.worksheets[0]
            for row in ws.iter_rows(values_only=True):
                row = list(row)
                while row and row[-1] is None:
                    del row[-1]
                if not row:
                    break
                yield tuple(x if x else '' for x in row)
        finally:
            wb.close()


class Excel(Tabular):
    read_handler = ExcelReadHandler
